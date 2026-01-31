import openpyxl
from pathlib import Path

study_dir = Path(r'data\Study 6 _CPID_Input Files - Anonymization')

print("="*80)
print("CHECKING ALL EXCEL FILES IN STUDY 6")
print("="*80)

for excel_file in study_dir.glob('*.xlsx'):
    print(f"\n\nFILE: {excel_file.name}")
    print("-"*80)
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print("\nFirst 3 rows:")
        for row_num in range(1, min(4, ws.max_row + 1)):
            print(f"  Row {row_num}:")
            row_values = []
            for col in range(1, min(11, ws.max_column + 1)):
                value = ws.cell(row_num, col).value
                if value is not None:
                    row_values.append(f"Col{col}={value}")
            print(f"    {', '.join(row_values)}")
        
        # Search for specific keywords
        keywords = ['Missing Visit', 'Missing Page', 'Input files', 'Total Queries']
        found_keywords = []
        for row_num in range(1, min(6, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                value = str(ws.cell(row_num, col).value or '')
                for keyword in keywords:
                    if keyword in value:
                        found_keywords.append(f"{keyword} at Row{row_num},Col{col}")
        
        if found_keywords:
            print("\nFOUND KEYWORDS:")
            for found in found_keywords:
                print(f"  ✓ {found}")
        
    except Exception as e:
        print(f"ERROR: {e}")

