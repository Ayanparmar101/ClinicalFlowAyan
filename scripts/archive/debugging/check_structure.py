import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook(r'data\Study 6 _CPID_Input Files - Anonymization\CPID_EDC_Metrics_URSV2.0_updated.xlsx')
ws = wb.active

print("="*80)
print("FIRST 15 COLUMNS - ROW BY ROW")
print("="*80)

for row_num in range(1, 8):
    print(f"\nROW {row_num}:")
    for col in range(1, 16):
        value = ws.cell(row_num, col).value
        print(f"  Col {col:2d}: {value}")

print("\n" + "="*80)
print("WHERE IS 'Input files - Missing Visits'?")
print("="*80)

# Search for this text in all cells
for row_num in range(1, 10):
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(row_num, col).value or '')
        if 'Missing Visit' in value or 'Missing Page' in value:
            print(f"Found at Row {row_num}, Col {col}: '{value}'")
