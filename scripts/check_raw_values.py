import openpyxl
import pandas as pd

# Load the Excel file
wb = openpyxl.load_workbook(r'data\Study 6 _CPID_Input Files - Anonymization\CPID_EDC_Metrics_URSV2.0_updated.xlsx')
ws = wb.active

# Get headers by combining 3 rows
headers = []
for col in range(1, ws.max_column + 1):
    h1 = str(ws.cell(1, col).value or '')
    h2 = str(ws.cell(2, col).value or '')
    h3 = str(ws.cell(3, col).value or '')
    combined = f"{h1} {h2} {h3}".strip()
    combined = ' '.join(combined.split())
    headers.append(combined)

print("="*80)
print("CHECKING COLUMN 8 (Missing Visits)")
print("="*80)
print(f"Column 8 header: {headers[7] if len(headers) > 7 else 'N/A'}")
print("\nFirst 10 data rows for column 8:")
for row in range(4, 14):
    value = ws.cell(row, 8).value
    print(f"Row {row}: {value} (type: {type(value).__name__})")

print("\n" + "="*80)
print("CHECKING COLUMN 9 (Missing Pages)")
print("="*80)
print(f"Column 9 header: {headers[8] if len(headers) > 8 else 'N/A'}")
print("\nFirst 10 data rows for column 9:")
for row in range(4, 14):
    value = ws.cell(row, 9).value
    print(f"Row {row}: {value} (type: {type(value).__name__})")

# Check if all values are 0
print("\n" + "="*80)
print("CHECKING ALL DATA ROWS")
print("="*80)

col8_values = []
col9_values = []
for row in range(4, ws.max_row + 1):
    col8_values.append(ws.cell(row, 8).value)
    col9_values.append(ws.cell(row, 9).value)

print(f"\nColumn 8 (Missing Visits) unique values: {set(col8_values)}")
print(f"Column 9 (Missing Pages) unique values: {set(col9_values)}")

print("\n" + "="*80)
print("CONCLUSION")
print("="*80)
if set(col8_values) == {0} or set(col8_values) == {0, None}:
    print("⚠️ ALL values in 'Input files - Missing Visits' column are 0")
if set(col9_values) == {0} or set(col9_values) == {0, None}:
    print("⚠️ ALL values in 'Missing Page' column are 0")
